#!/usr/bin/env python3
"""
LCA Quarterly Data Refresh Script

Downloads quarterly LCA disclosure data from DOL OFLC, extracts certified
H-1B employer names, and produces lca-sponsors.json with a rolling N-quarter
window. REBUILDS from scratch each run (no incremental merge) so the rolling
window is always exact — no stale employers from dropped quarters.

Usage:
  python3 lca-quarterly-refresh.py --auto                    # detect latest available quarter
  python3 lca-quarterly-refresh.py --quarter FY2026_Q2       # specify latest quarter
  python3 lca-quarterly-refresh.py --dry-run                 # show what would change
  python3 lca-quarterly-refresh.py --window 4                # use 4-quarter window

Output: lca-sponsors.json — name-only: {"_meta": {...}, "employers": [...]}
        Backward compatible with enrich-jobs.js loadLcaSponsors() (visa.js).

After generating the file, upload to R2:
  aws s3 cp lca-sponsors.json s3://$R2_BUCKET_NAME/data/lca-sponsors.json \\
    --endpoint-url $R2_ENDPOINT

DOL source: https://www.dol.gov/agencies/eta/foreign-labor/performance
File pattern: LCA_Disclosure_Data_FY{year}_Q{n}.xlsx
"""

import argparse
import datetime
import json
import re
import sys
import tempfile
from pathlib import Path

try:
    import openpyxl
    import requests
except ImportError as e:
    print(f"Missing dependency: {e}", file=sys.stderr)
    print("Install: pip install openpyxl requests", file=sys.stderr)
    sys.exit(1)

DOL_BASE_URL = "https://www.dol.gov/sites/dolgov/files/ETA/oflc/pdfs/LCA_Disclosure_Data_{quarter}.xlsx"
DEFAULT_WINDOW = 5  # quarters to keep
LCA_FILTER = {"CASE_STATUS": "Certified", "VISA_CLASS": "H-1B"}
EMPLOYER_COL = "EMPLOYER_NAME"


def download_quarter(quarter: str, dest: Path) -> bool:
    url = DOL_BASE_URL.format(quarter=quarter)
    print(f"Downloading {quarter} from DOL...")
    try:
        resp = requests.get(url, timeout=120, stream=True)
        if resp.status_code == 404:
            print(f"  NOT FOUND (404) — {quarter} data not yet published", file=sys.stderr)
            return False
        resp.raise_for_status()
        size_mb = int(resp.headers.get("content-length", 0)) / 1024 / 1024
        print(f"  Size: {size_mb:.1f} MB")
        with open(dest, "wb") as f:
            for chunk in resp.iter_content(chunk_size=1024 * 1024):
                f.write(chunk)
        print(f"  Downloaded to {dest}")
        return True
    except requests.RequestException as e:
        print(f"  Download failed: {e}", file=sys.stderr)
        return False


def extract_employers(xlsx_path: Path) -> set[str]:
    print(f"Extracting employers from {xlsx_path.name}...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {name: i for i, name in enumerate(header)}

    if EMPLOYER_COL not in col_idx:
        print(f"  ERROR: {EMPLOYER_COL} column not found. Columns: {header[:20]}", file=sys.stderr)
        wb.close()
        return set()

    case_idx = col_idx.get("CASE_STATUS")
    visa_idx = col_idx.get("VISA_CLASS")
    employer_idx = col_idx[EMPLOYER_COL]

    employers = set()
    total = 0
    filtered = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        if case_idx is not None and row[case_idx] != LCA_FILTER["CASE_STATUS"]:
            continue
        if visa_idx is not None and row[visa_idx] != LCA_FILTER["VISA_CLASS"]:
            continue
        name = row[employer_idx]
        if name and isinstance(name, str):
            employers.add(name.strip())
            filtered += 1

    wb.close()
    print(f"  Rows: {total:,}, Certified H-1B: {filtered:,}, Unique employers: {len(employers):,}")
    return employers


def normalize_employer_name(name: str) -> str:
    if not name:
        return ""
    n = name.lower().strip()
    n = re.sub(
        r"\b(inc|llc|ltd|corp|co|lp|llp|plc|gmbh|ag|sa|nv|bv|pte|pvt|limited|incorporated|corporation|company|group|holdings?|technologies?|solutions?|services?|systems?)\.?",
        "",
        n,
    )
    n = re.sub(r"[^a-z0-9\s]", " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n


def parse_quarter(q: str) -> tuple[int, int]:
    m = re.match(r"FY(\d{4})_Q(\d)", q)
    if not m:
        raise ValueError(f"Invalid quarter format: {q}. Expected FYXXXX_QN")
    return int(m[1]), int(m[2])


def next_quarter(q: str) -> str:
    fy, qn = parse_quarter(q)
    qn += 1
    if qn > 4:
        qn = 1
        fy += 1
    return f"FY{fy}_Q{qn}"


def build_quarter_window(latest: str, window: int) -> list[str]:
    quarters = [latest]
    fy, qn = parse_quarter(latest)
    for _ in range(window - 1):
        qn -= 1
        if qn < 1:
            qn = 4
            fy -= 1
        quarters.append(f"FY{fy}_Q{qn}")
    return quarters


def load_existing(path: Path) -> dict:
    if not path.exists():
        return {"_meta": {}, "employers": []}
    with open(path) as f:
        return json.load(f)


def detect_latest_available() -> str | None:
    """Auto-detect the latest quarter with DOL data published.
    DOL FY starts October: Q1=Oct-Dec, Q2=Jan-Mar, Q3=Apr-Jun, Q4=Jul-Sep.
    DOL publishes ~2-3 months after quarter end, so the current quarter
    is often not yet available — walks backwards until a 200 is found."""
    now = datetime.datetime.now()
    fiscal_month = (now.month - 10) % 12  # Oct=0, Nov=1, ... Sep=11
    qn = fiscal_month // 3 + 1
    fy = now.year + 1 if now.month >= 10 else now.year  # FY designation (not start year)

    for _ in range(6):  # Try up to 6 quarters back (DOL can lag ~6 months)
        candidate = f"FY{fy}_Q{qn}"
        url = DOL_BASE_URL.format(quarter=candidate)
        print(f"Auto-detect: checking {candidate}...")
        try:
            resp = requests.head(url, timeout=30, allow_redirects=True)
            if resp.status_code == 200:
                print(f"  {candidate} available")
                return candidate
            print(f"  {candidate}: HTTP {resp.status_code}")
        except requests.RequestException as e:
            print(f"  {candidate}: {e}", file=sys.stderr)
        qn -= 1
        if qn < 1:
            qn = 4
            fy -= 1

    return None


def validate_match_rate(old_employers: set[str], new_employers: set[str], threshold: float = 0.02) -> bool:
    if not old_employers:
        print("  No existing data to compare against — skipping validation")
        return True

    old_norm = {normalize_employer_name(e) for e in old_employers}
    new_norm = {normalize_employer_name(e) for e in new_employers}

    retained = old_norm & new_norm
    retention_rate = len(retained) / len(old_norm) if old_norm else 0

    print(f"  Previous: {len(old_employers):,} employers ({len(old_norm):,} normalized)")
    print(f"  New:      {len(new_employers):,} employers ({len(new_norm):,} normalized)")
    print(f"  Retained: {len(retained):,} ({retention_rate:.1%})")
    print(f"  Added:    {len(new_norm - old_norm):,}")
    print(f"  Removed:  {len(old_norm - new_norm):,}")

    if retention_rate < (1 - threshold):
        print(f"  WARNING: Retention rate {retention_rate:.1%} is below {(1-threshold):.0%} threshold", file=sys.stderr)
        print(f"  This may indicate a normalization regression or data format change", file=sys.stderr)
        return False

    return True


def main():
    parser = argparse.ArgumentParser(description="LCA Quarterly Data Refresh")
    parser.add_argument("--quarter", help="Latest quarter to include (e.g., FY2026_Q1)")
    parser.add_argument("--auto", action="store_true", help="Auto-detect latest available quarter")
    parser.add_argument("--window", type=int, default=DEFAULT_WINDOW, help=f"Rolling window size (default: {DEFAULT_WINDOW} quarters)")
    parser.add_argument("--output", default="lca-sponsors.json", help="Output file path")
    parser.add_argument("--existing", help="Path to existing lca-sponsors.json (for validation comparison only)")
    parser.add_argument("--dry-run", action="store_true", help="Show what would change without writing")
    args = parser.parse_args()

    if not args.quarter and not args.auto:
        parser.error("Specify --quarter or --auto")

    # Determine the latest quarter
    latest = args.quarter or detect_latest_available()
    if not latest:
        print("ERROR: No recent quarter data available from DOL", file=sys.stderr)
        sys.exit(1)

    # Build the rolling window of N quarters ending at latest
    quarters = build_quarter_window(latest, args.window)
    print(f"\n=== LCA Quarterly Refresh ===")
    print(f"Latest quarter: {latest}")
    print(f"Window ({args.window} quarters): {quarters}\n")

    # Download and extract employers from ALL quarters in the window.
    # Rebuild from scratch — no merge with existing. Guarantees the rolling
    # window is exact: employers from quarters outside the window are dropped.
    all_employers: set[str] = set()
    with tempfile.TemporaryDirectory() as tmpdir:
        for q in quarters:
            xlsx_path = Path(tmpdir) / f"LCA_Disclosure_Data_{q}.xlsx"
            if not download_quarter(q, xlsx_path):
                if q == latest:
                    print(f"\nERROR: Could not download latest quarter {q}", file=sys.stderr)
                    sys.exit(1)
                print(f"  WARNING: {q} unavailable — window may be smaller than {args.window}", file=sys.stderr)
                continue
            employers = extract_employers(xlsx_path)
            all_employers |= employers
            print(f"  Running total: {len(all_employers):,} unique employers\n")

    if not all_employers:
        print("ERROR: No employers extracted from any quarter", file=sys.stderr)
        sys.exit(1)

    # Load existing data for validation (comparison only — output is built from scratch)
    existing_path = Path(args.existing) if args.existing else Path(args.output)
    existing_data = load_existing(existing_path)
    existing_employers = set(existing_data.get("employers", []))
    existing_quarters = existing_data.get("_meta", {}).get("quarters", [])

    print(f"Existing data: {len(existing_employers):,} employers from {existing_quarters}")

    # Validate retention rate (sanity check — informational, not a hard gate)
    print("\nValidation:")
    if existing_employers:
        validate_match_rate(existing_employers, all_employers)

    # Build output
    output = {
        "_meta": {
            "description": "DOL OFLC LCA certified H-1B employer names (rolling window)",
            "source": "DOL OFLC LCA Disclosure Data",
            "quarters": quarters,
            "generated": datetime.datetime.now().isoformat()[:10],
            "filter": f"CASE_STATUS={LCA_FILTER['CASE_STATUS']}, VISA_CLASS={LCA_FILTER['VISA_CLASS']}",
            "total_employers": len(all_employers),
            "format": "name-only (backward compatible with enrich-jobs.js loadLcaSponsors)",
        },
        "employers": sorted(all_employers),
    }

    if args.dry_run:
        print(f"\n=== DRY RUN ===")
        print(f"Would write {len(all_employers):,} employers to {args.output}")
        print(f"Quarters: {quarters}")
        if existing_employers:
            delta_added = len(all_employers - existing_employers)
            delta_removed = len(existing_employers - all_employers)
            print(f"New employers: +{delta_added:,}")
            print(f"Removed employers: -{delta_removed:,}")
        print(f"Meta: {json.dumps(output['_meta'], indent=2)}")
        return

    # Write output
    out_path = Path(args.output)
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2)
    size_kb = out_path.stat().st_size / 1024
    print(f"\nWritten to {out_path} ({size_kb:.0f} KB)")

    print(f"\nNext step: upload to R2 (data/lca-sponsors.json)")
    print(f"  aws s3 cp {out_path} s3://$R2_BUCKET_NAME/data/lca-sponsors.json --endpoint-url $R2_ENDPOINT")


if __name__ == "__main__":
    main()
